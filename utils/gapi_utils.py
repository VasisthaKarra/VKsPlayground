#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#!pip install --upgrade google-api-python-client google-auth-httplib2 google-auth-oauthlib

from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
import io
from googleapiclient.http import MediaIoBaseDownload
#from apiclient.http import MediaFileUpload

import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from apiclient import errors, discovery
import mimetypes
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from openpyxl.utils import get_column_letter, column_index_from_string

def gmail():
    """Shows basic usage of the Gmail API.
    Lists the user's Gmail labels.
    """
    creds = None
    # If modifying these scopes, delete the file token.pickle.
    SCOPES = ['https://www.googleapis.com/auth/gmail.modify']
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    #if os.path.exists('gmailtoken.pickle'):
            #with open('gmailtoken.pickle', 'rb') as token:
            #creds = pickle.load(token)
    if os.path.exists(Path(__file__).resolve().parent.joinpath('support/gmailtoken.pickle')):
        creds = Credentials.from_authorized_user_file(Path(__file__).resolve().parent.joinpath('support/gmailtoken.pickle'), SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                Path(__file__).resolve().parent.joinpath('support/gAPI_cred.json'), SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        #with open('gmailtoken.pickle', 'wb') as token:
            #pickle.dump(creds, token)
        with open(Path(__file__).resolve().parent.joinpath('support/gmailtoken.pickle'), 'w') as token:
            token.write(creds.to_json())

    service = build('gmail', 'v1', credentials=creds)
    
    return service


def CreateMessage(sender, to, subject, cc,msgHtml):
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = to
    msg['Cc'] = cc
    #msg.attach(MIMEText(msgPlain, 'plain'))
    msg.attach(MIMEText(msgHtml, 'html'))
    raw = base64.urlsafe_b64encode(msg.as_bytes())
    raw = raw.decode()
    body = {'raw': raw}
    return body

def SendMessage(userId, message):
    try:
        gmail().users().messages().send(userId= "me", body=message).execute()
        #print ('Message Id: %s' % message['id'])
    except errors.HttpError or error:
        print ('An error occurred: %s' % error)

def create_message_with_attachment(
    sender, to, subject, message_text, file):
    """Create a message for an email.

      Args:
        sender: Email address of the sender.
        to: Email address of the receiver.
        subject: The subject of the email message.
        message_text: The text of the email message.
        file: The path to the file to be attached.

      Returns:
        An object containing a base64url encoded email object.
    """
    message = MIMEMultipart()
    message['to'] = to
    message['from'] = sender
    message['subject'] = subject

    msg = MIMEText(message_text)
    message.attach(msg)

    content_type, encoding = mimetypes.guess_type(file)

    if content_type is None or encoding is not None:
        content_type = 'application/octet-stream'
    main_type, sub_type = content_type.split('/', 1)
    if main_type == 'text':
        fp = open(file, 'rb')
        msg = MIMEText(fp.read(), _subtype=sub_type)
        fp.close()
    elif main_type == 'image':
        fp = open(file, 'rb')
        msg = MIMEImage(fp.read(), _subtype=sub_type)
        fp.close()
    elif main_type == 'audio':
        fp = open(file, 'rb')
        msg = MIMEAudio(fp.read(), _subtype=sub_type)
        fp.close()
    else:
        fp = open(file, 'rb')
        msg = MIMEBase(main_type, sub_type)
        msg.set_payload(fp.read())
        fp.close()
    filename = os.path.basename(file)
    msg.add_header('Content-Disposition', 'attachment', filename=filename)
    message.attach(msg)

    #return {'raw': base64.urlsafe_b64encode(message.as_string())}
    return {'raw': base64.urlsafe_b64encode(message.as_bytes())}





def gsheets():
    """Shows basic usage of the Sheets API.
    Prints values from a sample spreadsheet.
    """
    creds = None
    # If modifying these scopes, delete the file token.pickle.
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists(Path(__file__).resolve().parent.joinpath('support/gsheetstoken.pickle')):
        with open(Path(__file__).resolve().parent.joinpath('support/gsheetstoken.pickle'), 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                Path(__file__).resolve().parent.joinpath('support/gAPI_cred.json'), SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(Path(__file__).resolve().parent.joinpath('support/gsheetstoken.pickle'), 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)
    return service

from pathlib import Path

def gdrive():
    """Shows basic usage of the Drive v3 API.
    Prints the names and ids of the first 10 files the user has access to.
    """
    creds = None
    # If modifying these scopes, delete the file token.pickle.
    SCOPES = ['https://www.googleapis.com/auth/drive']
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists(Path(__file__).resolve().parent.joinpath('support/gdrivetoken.pickle')):
        with open(Path(__file__).resolve().parent.joinpath('support/gdrivetoken.pickle'), 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                Path(__file__).resolve().parent.joinpath('support/gAPI_cred.json'), SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(Path(__file__).resolve().parent.joinpath('support/gdrivetoken.pickle'), 'wb') as token:
            pickle.dump(creds, token)

    service = build('drive', 'v3', credentials=creds)

    # Call the Drive v3 API
    #my_function(service)
    return service

def searchDriveByQuery(query):
    
    page_token = None
    # Call the Drive v3 API
    results = gdrive().files().list(
        pageSize=10, q= query,
                   spaces='drive', supportsAllDrives= True, fields='files(id, name, mimeType,parents, webViewLink)', pageToken=page_token).execute()

    return results['files']

def searchDriveByFileId(fileId):
    
    # Call the Drive v3 API
    results = gdrive().files().get(
    fileId = fileId,
                   supportsAllDrives= True, fields='id, name, mimeType,parents, webViewLink').execute()

    return results
    
def downloadFileFromDrive(fileId):
    
    request = gdrive().files().get_media(fileId=fileId)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        
    return fh
        
def downloadGoogleDocFromDrive(fileId, mimeType):
    
    request = gdrive().files().export_media(fileId=fileId,
                                             mimeType=mimeType)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        
    return fh


import numpy as np
import pandas as pd

def gapi_write_to_sheet(df, sheet_id, sheet_name, mode='replace'):
    print(len(df))
    gapi_write_to_sheet_batch(df.iloc[:1000,:], sheet_id, sheet_name, mode)
    i = 1000
    while i<=len(df):
        gapi_write_to_sheet_batch(df.iloc[i:i+1000,:], sheet_id, sheet_name, 'append') 
        i = i+1000

def gapi_write_to_sheet_batch(add_df, sheet_id, sheet_name, mode = 'replace'):
    '''
    mode - replace/append    
    '''
    sheets_client = gsheets()
    if(mode == 'replace'):
        sheet = sheets_client.spreadsheets().values().get(
                                        spreadsheetId=sheet_id,
                                        range = sheet_name).execute()
        values = sheet.get('values', [])
        try:
            df = pd.DataFrame(values[1:], columns = values[0][:max([len(X) for X in values[1:]])])
            for col in values[0][:max([len(X) for X in values[1:]])]:
                df[col] = np.nan
            delete_set = np.zeros(shape=(df.shape[0]+1,df.shape[1]))
            delete_list = delete_set.tolist()
            delete_list = [['' for X in Y] for Y in delete_list]

            start_col = 'A'
            end_col = get_column_letter(df.shape[1])
            range_name = f"{sheet_name}!{start_col}1:{end_col}{str(len(delete_list))}"
            body = {'values': delete_list}

            result = sheets_client.spreadsheets().values().update(
                    spreadsheetId=sheet_id, 
                    range=range_name,
                    valueInputOption='USER_ENTERED', 
                    body=body).execute()
        except IndexError:
            a = f"No records in {sheet_name}"
            print(a)
        
    headers = [list(add_df.columns)]
    values = add_df.fillna('').astype(str).values.tolist()
    
    if(mode == 'replace'):
        start_col = 'A'
        end_col = get_column_letter(len(headers[0]))
        range_name = f"{sheet_name}!{start_col}1:{end_col}1"
        body = {'values': headers}

        result = sheets_client.spreadsheets().values().update(
            spreadsheetId=sheet_id, 
            range=range_name,
            valueInputOption='USER_ENTERED', 
            body=body).execute()

        range_name = f"{sheet_name}!{start_col}2:{end_col}{str(len(values)+1)}"
        body = {'values': values}

        result = sheets_client.spreadsheets().values().update(
            spreadsheetId=sheet_id, 
            range=range_name,
            valueInputOption='USER_ENTERED', 
            body=body).execute()
        
    elif(mode == 'append'):
        body = {'values': values}
        result = sheets_client.spreadsheets().values().append(
            spreadsheetId=sheet_id, 
            range=sheet_name,
            valueInputOption='USER_ENTERED', 
            body=body).execute()
    else:
        print("Invalid mode")

sheets_client = gsheets()
drive_client = gdrive()

def gapi_read_sheet(sheet_id, sheet_name, value_render_option='FORMATTED_VALUE'):
    #sheets_client = gsheets()
    sheet = sheets_client.spreadsheets().values().get(
                                spreadsheetId= sheet_id,
                                range=sheet_name,
                                valueRenderOption=value_render_option).execute()
    values = sheet.get('values', [])
    values = [X for X in values if len(X)!=0]

    df = pd.DataFrame(values[1:], columns = values[0][:max([len(X) for X in values[1:]])])
    for col in list(set(values[0]).difference(set(values[0][:max([len(X) for X in values[1:]])]))):
        df[col] = np.nan
    return df

def gapi_write_to_row(sheet_id, sheet_name, insert_values, index, col_name):
    #sheets_client = gsheets()
    sheet = sheets_client.spreadsheets().values().get(
                                            spreadsheetId=sheet_id,
                                            range = f'{sheet_name}').execute()
    values = sheet.get('values', [])
    columns = values[0]
    
    #Identifying the values
    if((isinstance(insert_values[0],str))|(isinstance(insert_values[0],int))|(isinstance(insert_values[0],float))):
        #One dimentional
        n_cols = len(insert_values)
        body = {'values': [insert_values]}
    else:
        n_cols = max([len(X) for X in insert_values])
        #length = max(map(len, insert_values))
        #insert_values=[X+['']*(length-len(X)) for X in insert_values]
        body = {'values': insert_values}
    
    ##Calculating columns
    col_index = columns.index(col_name)
    start_col = get_column_letter(col_index+1)
    end_col = get_column_letter(col_index+1+(n_cols-1))
    
    start_row = index+2
    end_row = start_row+len(insert_values)-1

    range_name = f"{sheet_name}!{start_col}{start_row}:{end_col}{end_row}"
    value_input_option = 'USER_ENTERED'
    
    result = sheets_client.spreadsheets().values().update(
        spreadsheetId=sheet_id, range=range_name,
        valueInputOption=value_input_option, body=body).execute()
    

    
def gapi_write_to_row2(sheet_id, sheet_name, insert_values, index, col_name,sheet=None):
    if(sheet==None):
        sheets_client = gsheets()
        sheet = sheets_client.spreadsheets().values().get(spreadsheetId=sheet_id,range = f'{sheet_name}').execute()
    
    values = sheet.get('values', [])
    columns = values[0]
    
    #Identifying the values
    if((isinstance(insert_values[0],str))|(isinstance(insert_values[0],int))|(isinstance(insert_values[0],float))):
        #One dimentional
        n_cols = len(insert_values)
        body = {'values': [insert_values]}
    else:
        n_cols = max([len(X) for X in insert_values])
        #length = max(map(len, insert_values))
        #insert_values=[X+['']*(length-len(X)) for X in insert_values]
        body = {'values': insert_values}
    
    ##Calculating columns
    col_index = columns.index(col_name)
    start_col = get_column_letter(col_index+1)
    end_col = get_column_letter(col_index+1+(n_cols-1))
    
    start_row = index+2
    end_row = start_row+len(insert_values)-1

    range_name = f"{sheet_name}!{start_col}{start_row}:{end_col}{end_row}"
    value_input_option = 'USER_ENTERED'
    
    result = sheets_client.spreadsheets().values().update(
        spreadsheetId=sheet_id, range=range_name,
        valueInputOption=value_input_option, body=body).execute()
    
    
def create_spreadsheet(title, primary_sheet_title, parent_folder_id = '1L5z'):

    def create(title, sheet_title):
        spreadsheet_body = {
            'properties': {
                'title': title
            },
            'sheets': [
                {
                    'properties': {
                    'sheetId': 0,
                    'title': sheet_title,
                    'index': 0,
                    'sheetType': 'GRID',
                    'gridProperties': {
                        'rowCount': 1000, 
                        'columnCount': 26
                    }
                    }
                }
            ]
        }
        request = sheets_client.spreadsheets().create(body=spreadsheet_body)
        response = request.execute()

        return response['spreadsheetId']

    def move(file_id, folder_id):
        # Retrieve the existing parents to remove
        file = drive_client.files().get(fileId=file_id,
                                         fields='parents').execute()
        previous_parents = ",".join(file.get('parents'))
        # Move the file to the new folder
        file = drive_client.files().update(fileId=file_id,
                                            addParents=folder_id,
                                            removeParents=previous_parents,
                                            fields='id, parents').execute()

        return True
    
    spreadsheet_id = create(title, primary_sheet_title)
    move_status = move(spreadsheet_id, parent_folder_id)
    if(move_status):
        return spreadsheet_id
    else:
        return None
    
def add_new_sheet(spreadsheet_id, sheetId, sheet_title):
    batch_update_spreadsheet_request_body = {
        'requests': [{
                      "addSheet": {
                        "properties": {
                          "sheetId": sheetId,
                          "title": sheet_title,
                          "index": 1,
                          "sheetType": "GRID",
                          "gridProperties": {
                            "rowCount": 1000,
                            "columnCount": 26
                          }
                        }
                      }
                    }]
    }

    request = sheets_client.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=batch_update_spreadsheet_request_body)
    response = request.execute()
    return True

def update_existing_sheet(spreadsheet_id, sheetId, sheet_title):
    batch_update_spreadsheet_request_body = {
    'requests': [{
                  "updateSheetProperties": {
                    "properties": {
                      "sheetId": sheetId,
                      "title": sheet_title
                    },
                    "fields": "title"
                  }
                }]
    }

    request = sheets_client.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=batch_update_spreadsheet_request_body)
    response = request.execute()
    return True